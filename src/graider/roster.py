"""Roster parsing: CSV/XLSX -> validated Students -> Groups."""

from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook
from pydantic import ValidationError

from graider.errors import RosterError
from graider.models import Group, Student

# Canonical field -> accepted header names (after normalization: lowercased,
# runs of space/underscore/hyphen collapsed to a single space).
EMAIL_HEADERS = {"email", "e mail", "mail", "student email", "student e mail"}
GROUP_HEADERS = {"group", "group number", "groupnumber", "group no", "team", "team number"}
NAME_HEADERS = {"name", "student", "student name", "full name"}


def read_roster(path: Path) -> list[Student]:
    """Parse a roster file into validated Students. Raises RosterError."""
    if not path.exists():
        raise RosterError(f"Roster file not found: {path}")

    headers, raw_rows = _load_rows(path)
    field_by_col = _map_headers(headers, path)

    students: list[Student] = []
    errors: list[str] = []
    seen: dict[str, int] = {}

    for offset, raw in enumerate(raw_rows):
        rownum = offset + 2  # header is row 1
        row = {field: raw[col] if col < len(raw) else "" for col, field in field_by_col.items()}
        if not any(v.strip() for v in row.values()):
            continue  # blank row

        try:
            student = Student(
                email=row.get("email", ""),
                group_number=row.get("group_number", ""),
                name=row.get("name") or None,
            )
        except ValidationError as exc:
            for err in exc.errors():
                errors.append(f"row {rownum}: {err['msg']}")
            continue

        if student.email in seen:
            errors.append(
                f"row {rownum}: duplicate student {student.email} "
                f"(first seen row {seen[student.email]})"
            )
            continue
        seen[student.email] = rownum
        students.append(student)

    if errors:
        raise RosterError("Roster has problems:\n  " + "\n  ".join(errors))
    if not students:
        raise RosterError(f"No students found in {path}")
    return students


def group_students(students: list[Student]) -> list[Group]:
    """Aggregate students into groups, ordered by first appearance."""
    buckets: dict[str, list[Student]] = {}
    for student in students:
        buckets.setdefault(student.group_number, []).append(student)
    return [Group(number=number, members=members) for number, members in buckets.items()]


# --- internals ---------------------------------------------------------------


def _load_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv(path)
    elif suffix in (".xlsx", ".xlsm"):
        rows = _read_xlsx(path)
    else:
        raise RosterError(f"Unsupported roster format {path.suffix!r} (use .csv or .xlsx)")
    if not rows:
        raise RosterError(f"{path} is empty")
    return rows[0], rows[1:]


def _read_csv(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8-sig") as fh:
        return [[(cell or "").strip() for cell in row] for row in csv.reader(fh)]


def _read_xlsx(path: Path) -> list[list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        return [[_cell_to_str(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))  # 1.0 -> "1"
    return str(value).strip()


def _map_headers(headers: list[str], path: Path) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for col, header in enumerate(headers):
        field = _canonical_field(header)
        if field is not None:
            mapping[col] = field
    fields = set(mapping.values())
    if "email" not in fields:
        raise RosterError(f"{path}: no email column found (looked for {sorted(EMAIL_HEADERS)})")
    if "group_number" not in fields:
        raise RosterError(f"{path}: no group column found (looked for {sorted(GROUP_HEADERS)})")
    return mapping


def _canonical_field(header: str) -> str | None:
    key = re.sub(r"[\s_\-]+", " ", (header or "").strip().lower()).strip()
    if key in EMAIL_HEADERS:
        return "email"
    if key in GROUP_HEADERS:
        return "group_number"
    if key in NAME_HEADERS:
        return "name"
    return None
